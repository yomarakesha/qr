import io
import os
import re
import sqlite3
import time
from functools import wraps
from flask import Flask, g, jsonify, request, send_file, render_template, abort, session, redirect, url_for
from werkzeug.security import check_password_hash

import qrcode
from qrcode.constants import ERROR_CORRECT_L, ERROR_CORRECT_M, ERROR_CORRECT_Q, ERROR_CORRECT_H
from PIL import Image

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "contacts.db"))

app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")

ECL_MAP = {"L": ERROR_CORRECT_L, "M": ERROR_CORRECT_M, "Q": ERROR_CORRECT_Q, "H": ERROR_CORRECT_H}


def get_db():
    db = getattr(g, "_db", None)
    if db is None:
        db = g._db = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
        # Unicode-aware lowercasing for case-insensitive search (SQLite LOWER is ASCII-only)
        db.create_function("pylower", 1, lambda s: (s or "").lower())
    return db


@app.teardown_appcontext
def close_db(exc):
    db = getattr(g, "_db", None)
    if db is not None:
        db.close()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name   TEXT NOT NULL DEFAULT '',
            last_name   TEXT NOT NULL DEFAULT '',
            first_name  TEXT NOT NULL DEFAULT '',
            title       TEXT DEFAULT '',
            department  TEXT DEFAULT '',
            phone       TEXT DEFAULT '',
            email       TEXT DEFAULT '',
            file_name   TEXT DEFAULT '',
            size        INTEGER DEFAULT 600,
            ecl         TEXT DEFAULT 'M',
            created_at  INTEGER NOT NULL,
            updated_at  INTEGER NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at    INTEGER NOT NULL
        )
    """)
    cols = [r[1] for r in conn.execute("PRAGMA table_info(contacts)").fetchall()]
    if "full_name" not in cols:
        conn.execute("ALTER TABLE contacts ADD COLUMN full_name TEXT NOT NULL DEFAULT ''")
    if "qr_mode" not in cols:
        conn.execute("ALTER TABLE contacts ADD COLUMN qr_mode TEXT NOT NULL DEFAULT 'text'")
    if "email" not in cols:
        conn.execute("ALTER TABLE contacts ADD COLUMN email TEXT DEFAULT ''")
    conn.execute("""
        UPDATE contacts
           SET full_name = TRIM(COALESCE(first_name,'') || ' ' || COALESCE(last_name,''))
         WHERE (full_name IS NULL OR full_name = '')
    """)
    conn.commit()
    conn.close()


PHONE_RE = re.compile(r"^\+?[0-9 ()\-]+$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
QR_MODES = {"vcard", "text", "tel"}


def validate(data, partial=False):
    errors = {}
    if not partial or "fullName" in data:
        if not (data.get("fullName") or "").strip():
            errors["fullName"] = "Ady hökmany"
    phone = (data.get("phone") or "").strip()
    if phone:
        if not PHONE_RE.match(phone) or len(re.sub(r"\D", "", phone)) < 6:
            errors["phone"] = "Nädogry telefon"
    email = (data.get("email") or "").strip()
    if email and not EMAIL_RE.match(email):
        errors["email"] = "Nädogry email"
    ecl = (data.get("ecl") or "M").upper()
    if ecl not in ECL_MAP:
        errors["ecl"] = "Düzediş derejesi L/M/Q/H bolmaly"
    mode = (data.get("qrMode") or "text").lower()
    if mode not in QR_MODES:
        errors["qrMode"] = "QR mode: vcard / text / tel"
    if mode == "tel" and not (data.get("phone") or "").strip():
        errors["phone"] = "Tel rejimi üçin telefon hökmany"
    size = data.get("size", 600)
    try:
        size_i = int(size)
        if size_i < 100 or size_i > 4000:
            errors["size"] = "Ölçeg 100..4000 bolmaly"
    except (TypeError, ValueError):
        errors["size"] = "Ölçeg san bolmaly"
    return errors


def row_to_dict(row):
    return {
        "id": row["id"],
        "fullName": row["full_name"] if "full_name" in row.keys() else "",
        "title": row["title"],
        "department": row["department"],
        "phone": row["phone"],
        "email": row["email"] if "email" in row.keys() else "",
        "fileName": row["file_name"],
        "size": row["size"],
        "ecl": row["ecl"],
        "qrMode": (row["qr_mode"] if "qr_mode" in row.keys() else "text") or "text",
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }


def escape_vcard(s):
    return (s or "").replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def clean_phone(s):
    return re.sub(r"[\s()\-]", "", s or "")


def build_vcard(c):
    lines = ["BEGIN:VCARD", "VERSION:3.0"]
    name = (c.get("fullName") or "").strip()
    lines.append(f"N;CHARSET=UTF-8:{escape_vcard(name)};;;;")
    lines.append(f"FN;CHARSET=UTF-8:{escape_vcard(name)}")
    if c.get("department"):
        lines.append(f"ORG;CHARSET=UTF-8:{escape_vcard(c['department'])}")
    if c.get("title"):
        lines.append(f"TITLE;CHARSET=UTF-8:{escape_vcard(c['title'])}")
    if c.get("phone"):
        lines.append(f"TEL;TYPE=CELL:{clean_phone(c['phone'])}")
    if c.get("email"):
        lines.append(f"EMAIL;TYPE=INTERNET:{escape_vcard(c['email'])}")
    lines.append("END:VCARD")
    return "\r\n".join(lines)


def build_text(c):
    parts = []
    if c.get("department"): parts.append(f"Bölüm: {c['department']}")
    if c.get("title"):      parts.append(f"Wezipe: {c['title']}")
    if c.get("fullName"):   parts.append(f"Ady: {c['fullName']}")
    if c.get("phone"):      parts.append(f"Tel.: {c['phone']}")
    if c.get("email"):      parts.append(f"Email: {c['email']}")
    return "\n".join(parts)


def build_payload(c):
    mode = (c.get("qrMode") or "text").lower()
    if mode == "vcard":
        return build_vcard(c)
    if mode == "tel":
        return "tel:" + clean_phone(c.get("phone") or "")
    return build_text(c)


# ---------- Auth ----------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "unauthorized"}), 401
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        db = get_db()
        row = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        if row and check_password_hash(row["password_hash"], password):
            session["user"] = username
            nxt = request.args.get("next") or url_for("index")
            return redirect(nxt)
        error = "Ulanyjy ady ýa-da parol nädogry"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------- Routes ----------
@app.route("/")
@login_required
def index():
    return render_template("index.html", username=session.get("user"))


SORT_COLUMNS = {
    "updated": "updated_at",
    "created": "created_at",
    "name": "full_name",
    "department": "department",
}


def contacts_filters(args):
    """Build a WHERE clause + params from query args (q / department / qrMode)."""
    where, params = [], []
    q = (args.get("q") or "").strip().lower()
    if q:
        like = f"%{q}%"
        where.append(
            "(pylower(full_name) LIKE ? OR pylower(department) LIKE ? "
            "OR pylower(title) LIKE ? OR pylower(phone) LIKE ? OR pylower(email) LIKE ?)"
        )
        params += [like, like, like, like, like]
    department = (args.get("department") or "").strip()
    if department:
        where.append("department = ?")
        params.append(department)
    mode = (args.get("qrMode") or "").strip().lower()
    if mode in QR_MODES:
        where.append("COALESCE(qr_mode, 'text') = ?")
        params.append(mode)
    where_sql = (" WHERE " + " AND ".join(where)) if where else ""
    return where_sql, params


@app.get("/api/contacts")
@login_required
def list_contacts():
    args = request.args
    where_sql, params = contacts_filters(args)

    sort_col = SORT_COLUMNS.get((args.get("sort") or "updated").lower(), "updated_at")
    direction = "ASC" if (args.get("dir") or "desc").lower() == "asc" else "DESC"

    try:
        page = max(1, int(args.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = int(args.get("pageSize", 20))
    except (TypeError, ValueError):
        page_size = 20
    page_size = min(max(page_size, 1), 100)

    db = get_db()
    total = db.execute(f"SELECT COUNT(*) FROM contacts{where_sql}", params).fetchone()[0]
    pages = max(1, (total + page_size - 1) // page_size)
    page = min(page, pages)
    offset = (page - 1) * page_size

    rows = db.execute(
        f"SELECT * FROM contacts{where_sql} "
        f"ORDER BY {sort_col} COLLATE NOCASE {direction}, id DESC LIMIT ? OFFSET ?",
        params + [page_size, offset],
    ).fetchall()

    return jsonify({
        "items": [row_to_dict(r) for r in rows],
        "total": total,
        "page": page,
        "pageSize": page_size,
        "pages": pages,
    })


@app.get("/api/departments")
@login_required
def list_departments():
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT department FROM contacts "
        "WHERE TRIM(COALESCE(department, '')) <> '' ORDER BY department COLLATE NOCASE"
    ).fetchall()
    return jsonify([r[0] for r in rows])


@app.get("/api/contacts/<int:cid>")
@login_required
def get_contact(cid):
    db = get_db()
    row = db.execute("SELECT * FROM contacts WHERE id = ?", (cid,)).fetchone()
    if not row:
        abort(404)
    return jsonify(row_to_dict(row))


@app.post("/api/contacts")
@login_required
def create_contact():
    data = request.get_json(silent=True) or {}
    errors = validate(data)
    if errors:
        return jsonify({"errors": errors}), 400
    now = int(time.time() * 1000)
    db = get_db()
    cur = db.execute(
        """INSERT INTO contacts
           (full_name, title, department, phone, email, file_name, size, ecl, qr_mode, created_at, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            (data.get("fullName") or "").strip(),
            (data.get("title") or "").strip(),
            (data.get("department") or "").strip(),
            (data.get("phone") or "").strip(),
            (data.get("email") or "").strip(),
            (data.get("fileName") or "").strip(),
            int(data.get("size") or 600),
            (data.get("ecl") or "M").upper(),
            (data.get("qrMode") or "text").lower(),
            now, now,
        ),
    )
    db.commit()
    row = db.execute("SELECT * FROM contacts WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(row_to_dict(row)), 201


@app.put("/api/contacts/<int:cid>")
@login_required
def update_contact(cid):
    data = request.get_json(silent=True) or {}
    errors = validate(data)
    if errors:
        return jsonify({"errors": errors}), 400
    db = get_db()
    row = db.execute("SELECT * FROM contacts WHERE id = ?", (cid,)).fetchone()
    if not row:
        abort(404)
    now = int(time.time() * 1000)
    db.execute(
        """UPDATE contacts SET
           full_name=?, title=?, department=?,
           phone=?, email=?, file_name=?, size=?, ecl=?, qr_mode=?, updated_at=?
           WHERE id=?""",
        (
            (data.get("fullName") or "").strip(),
            (data.get("title") or "").strip(),
            (data.get("department") or "").strip(),
            (data.get("phone") or "").strip(),
            (data.get("email") or "").strip(),
            (data.get("fileName") or "").strip(),
            int(data.get("size") or 600),
            (data.get("ecl") or "M").upper(),
            (data.get("qrMode") or "text").lower(),
            now, cid,
        ),
    )
    db.commit()
    row = db.execute("SELECT * FROM contacts WHERE id = ?", (cid,)).fetchone()
    return jsonify(row_to_dict(row))


@app.delete("/api/contacts/<int:cid>")
@login_required
def delete_contact(cid):
    db = get_db()
    cur = db.execute("DELETE FROM contacts WHERE id = ?", (cid,))
    db.commit()
    if cur.rowcount == 0:
        abort(404)
    return "", 204


@app.get("/api/contacts/<int:cid>/vcard")
@login_required
def get_vcard(cid):
    db = get_db()
    row = db.execute("SELECT * FROM contacts WHERE id = ?", (cid,)).fetchone()
    if not row:
        abort(404)
    c = row_to_dict(row)
    payload = build_payload(c)
    mime = "text/vcard; charset=utf-8" if c.get("qrMode") == "vcard" else "text/plain; charset=utf-8"
    return payload, 200, {"Content-Type": mime}


def render_qr_png(text, size=600, ecl="M"):
    qr = qrcode.QRCode(
        version=None,
        error_correction=ECL_MAP.get(ecl.upper(), ERROR_CORRECT_M),
        box_size=10,
        border=2,
    )
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    img = img.resize((size, size), Image.NEAREST)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


@app.get("/api/contacts/<int:cid>/qr")
@login_required
def get_qr(cid):
    db = get_db()
    row = db.execute("SELECT * FROM contacts WHERE id = ?", (cid,)).fetchone()
    if not row:
        abort(404)
    c = row_to_dict(row)
    size = int(request.args.get("size", c["size"] or 600))
    ecl = request.args.get("ecl", c["ecl"] or "M").upper()
    download = request.args.get("download") == "1"
    payload = build_payload(c)
    buf = render_qr_png(payload, size=size, ecl=ecl)
    filename = (c["fileName"] or f"contact_{cid}") + ".png"
    return send_file(
        buf,
        mimetype="image/png",
        as_attachment=download,
        download_name=filename,
    )


@app.post("/api/qr/preview")
@login_required
def preview_qr():
    data = request.get_json(silent=True) or {}
    errors = validate(data)
    if errors:
        return jsonify({"errors": errors}), 400
    size = int(data.get("size") or 600)
    ecl = (data.get("ecl") or "M").upper()
    payload = build_payload(data)
    buf = render_qr_png(payload, size=size, ecl=ecl)
    return send_file(buf, mimetype="image/png")


@app.post("/api/vcard/preview")
@login_required
def preview_vcard():
    data = request.get_json(silent=True) or {}
    errors = validate(data)
    if errors:
        return jsonify({"errors": errors}), 400
    return jsonify({"vcard": build_payload(data), "mode": (data.get("qrMode") or "text").lower()})


# ---------- Excel import / export ----------
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Column order used for both export and the import template.
EXPORT_HEADERS = [
    "Bölüm", "Wezipe", "Ady we familiýasy", "Telefon", "Email",
    "QR rejimi", "Ölçeg", "ECL", "Faýl ady",
]
EXPORT_KEYS = [
    "department", "title", "fullName", "phone", "email",
    "qrMode", "size", "ecl", "fileName",
]
COL_WIDTHS = [16, 16, 24, 18, 28, 12, 9, 8, 22]

# Header text -> internal key (accepts Turkmen / English / Russian aliases, case-insensitive).
HEADER_ALIASES = {
    "department": ["bölüm", "bolum", "bölüm ady", "bolum ady", "department", "отдел"],
    "title":      ["wezipe", "title", "должность", "lavazym"],
    "fullName":   ["ady we familiýasy", "ady we familiyasy", "ady", "full name",
                   "fullname", "fio", "имя", "ф.и.о."],
    "phone":      ["telefon", "phone", "tel", "телефон"],
    "email":      ["email", "e-mail", "почта", "эл. почта"],
    "qrMode":     ["qr rejimi", "qr mode", "qrmode", "mode", "rejim", "режим"],
    "size":       ["ölçeg", "olceg", "size", "размер"],
    "ecl":        ["ecl", "düzediş", "duzedis", "коррекция"],
    "fileName":   ["faýl ady", "fayl ady", "faýl", "fayl", "file name", "filename", "файл"],
}


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="2563EB")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 22
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def workbook_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype=EXCEL_MIME, as_attachment=True, download_name=filename)


@app.get("/api/contacts/export.xlsx")
@login_required
def export_contacts():
    where_sql, params = contacts_filters(request.args)
    db = get_db()
    rows = db.execute(
        f"SELECT * FROM contacts{where_sql} ORDER BY full_name COLLATE NOCASE", params
    ).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontaktlar"
    ws.append(EXPORT_HEADERS)
    for r in rows:
        c = row_to_dict(r)
        ws.append([c.get(k, "") for k in EXPORT_KEYS])
    style_sheet(ws)
    return workbook_response(wb, "kontaktlar.xlsx")


@app.get("/api/contacts/template.xlsx")
@login_required
def template_contacts():
    wb = Workbook()
    ws = wb.active
    ws.title = "Şablon"
    ws.append(EXPORT_HEADERS)
    ws.append(["PÜweMHB", "başlygy", "M.Gasanow", "+993 12 93-01-66",
               "m.gasanow@example.com", "vcard", 600, "M", "M_Gasanow_qr"])
    ws.append(["Buhgalteriýa", "hasapçy", "A.Myradow", "+993 12 45-67-89",
               "a.myradow@example.com", "text", 600, "M", ""])
    style_sheet(ws)
    # Sample rows in a softer style so they read as examples, not real data.
    sample_font = Font(italic=True, color="64748B")
    for row in ws.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.font = sample_font
    return workbook_response(wb, "kontaktlar_shablon.xlsx")


def map_header(header_row):
    """Map column index -> internal key based on a header row."""
    lookup = {}
    for key, aliases in HEADER_ALIASES.items():
        for a in aliases:
            lookup[a] = key
    colmap = {}
    for idx, cell in enumerate(header_row or []):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        if name in lookup:
            colmap[idx] = lookup[name]
    return colmap


@app.post("/api/contacts/import")
@login_required
def import_contacts():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Faýl tapylmady"}), 400
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return jsonify({"error": "Excel faýly okap bolmady (.xlsx gerek)"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"created": 0, "failed": 0, "errors": []})

    colmap = map_header(rows[0])
    if not colmap:
        return jsonify({"error": "Sütün atlary tapylmady. Şablony ulanyň."}), 400

    db = get_db()
    created = 0
    errors = []
    for line_no, raw in enumerate(rows[1:], start=2):
        data = {}
        for idx, key in colmap.items():
            if idx >= len(raw):
                continue
            val = raw[idx]
            if val is None:
                continue
            if isinstance(val, float) and val.is_integer():
                val = int(val)
            sval = str(val).strip()
            if sval:
                data[key] = sval
        if not data:
            continue  # fully empty row
        errs = validate(data)
        if errs:
            errors.append({"row": line_no, "errors": errs})
            continue
        now = int(time.time() * 1000)
        db.execute(
            """INSERT INTO contacts
               (full_name, title, department, phone, email, file_name, size, ecl, qr_mode, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                (data.get("fullName") or "").strip(),
                (data.get("title") or "").strip(),
                (data.get("department") or "").strip(),
                (data.get("phone") or "").strip(),
                (data.get("email") or "").strip(),
                (data.get("fileName") or "").strip(),
                int(data.get("size") or 600),
                (data.get("ecl") or "M").upper(),
                (data.get("qrMode") or "text").lower(),
                now, now,
            ),
        )
        created += 1
    db.commit()
    return jsonify({"created": created, "failed": len(errors), "errors": errors[:50]})


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
